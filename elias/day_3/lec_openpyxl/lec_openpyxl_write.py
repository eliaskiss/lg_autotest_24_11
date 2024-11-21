from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from datetime import datetime, timedelta

###############################################################
# Create new Excel Workbook
###############################################################
wb = Workbook()

file_name = './new_excel.xlsx'

###############################################################
# Get Activate Worksheet
###############################################################
ws = wb.active # Sheet1

###############################################################
# Set Sheet Name
###############################################################
ws.title = 'Basic'

###############################################################
# Set Cell Value
###############################################################
ws['A1'] = 'Hello World'
ws['B1'] = 10
ws['C1'] = 20

###############################################################
# Use Formular
###############################################################
ws['D1'] = '=SUM(B1+C1)'

###############################################################
# Insert Comment into cell
###############################################################
comment = Comment('This is comment', 'Elias Kim', 100, 100)
ws['A1'].comment = comment

###############################################################
# Insert Rows
###############################################################
header = ['V1', 'V2', 'V3', 'V4', 'V5']
ws.append(header)
row = [10, 20, 30, 40, 50]
for _ in range(10):
    ws.append(row)

###############################################################
# Insert Datetime
###############################################################
ws['A14'] = datetime.now()
# ws['A14'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
ws['B14'] = datetime.now() + timedelta(days=1) # 내일
ws['C14'] = datetime.now() - timedelta(days=1) # 어제
ws['D14'] = datetime.now() + timedelta(hours=1) # 1시간 후
ws['E14'] = '2024-11-21'

###############################################################
# Merge/Unmerge Target Cell
###############################################################
ws['A16'] = 'Hello'
ws['B16'] = 'World'

# Way I
# ws.merge_cells(range_string='A16:B16')
# ws.unmerge_cells('A16:B16')

# Way II
ws.merge_cells(start_row=16, start_column=1, end_row=16, end_column=2)
ws.unmerge_cells(start_row=16, start_column=1, end_row=16, end_column=2)

###############################################################
# Insert Image
###############################################################
img = Image('buz.jpg')
ws.add_image(img, 'G1')

###############################################################
# Create 2nd Sheet
###############################################################
wb.create_sheet('Insert Delete Move')
wb.active = 1

################################################
# Select New Worksheet
################################################
# ws = wb.active
ws = wb['Insert Delete Move']

################################################
# Create Row Data
################################################
for i in range(20):
    # 'A':42 --> ord('A') + 1 --> 42 + 1 = 43, chr(43) --> B
    new_row = [f'{chr(ord("A") + j)}{i + 1}' for j in range(20)]  # A1, B1, C1, D1, ....
    ws.append(new_row)

################################################
# Insert Empty Row
################################################
# ws.insert_rows(2)

################################################
# Insert Empty Column
################################################
# ws.insert_cols(2)

################################################
# Delete Rows
################################################
# ws.delete_rows(5, 7) # From 5, Step 7 --> 5번재 Row부터 7개를 삭제

################################################
# Delete Columns
################################################
# ws.delete_cols(5, 7) # From 5, Step 7 --> 5번재 Column부터 7개를 삭제

################################################
# Move Range
################################################
# ws.move_range('A1:T1', rows=21, cols=1)

###############################################################
# Create Chart Sheet
# Doc Site: https://openpyxl.readthedocs.io/en/stable/charts/introduction.html
###############################################################
wb.create_sheet('Chart')
wb.active = 2
ws = wb['Chart']









###############################################################
# Save Excel File
###############################################################
wb.save(file_name)





























